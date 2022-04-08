from openpyxl import Workbook as wb
from openpyxl import load_workbook as lwb

file_source = "Книга10.xlsx"
file_result = file_source.replace(".", "_result.")

wb_source = lwb(file_source)
ws_source = wb_source.active

wb_result = wb()
ws_result = wb_result.active
k_wb_result = 0

end_source = ws_source.max_row + 1

for i in range(1, end_source):
    number_rows = int(ws_source.cell(i, 2).value)
    address = str(ws_source.cell(i, 1).value)

    for j in range(1, number_rows + 1):
        num_row_result = j + k_wb_result
        ws_result.cell(num_row_result, 1).value = address
        ws_result.cell(num_row_result, 2).value = j
        ws_result.cell(num_row_result, 3).value = str(ws_result.cell(num_row_result, 1).value) + " , " + str(ws_result.cell(num_row_result, 2).value)
        ws_result.cell(num_row_result, 4).value = i

    k_wb_result = k_wb_result + number_rows

wb_result.save(file_result)
